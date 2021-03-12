from openpyxl import load_workbook
import input_function

workbook = load_workbook(filename="sample.xlsx")
workbook.sheetnames
spreadsheet = workbook.active

print(">>>PYTHON PHONE DIRECTORY<<<")
entry = int(input("Please select your desired entry:\n1. Creat New Contact\n2. Edit\n3. Remove\n4. Display Records\n5. Exit\n"))

def input_function(entry):
    return entry
input_function(entry)

record = []
while True:
    
    if entry == 1:
        print("Creating New contact record")
        spreadsheet["a2"] = "Shahid"
        spreadsheet["b2"] = "03014949223"
        save = workbook.save(filename = "sample.xlsx")
        record.append(save)
        break
    elif entry == 2:
        pass

    elif entry == 3:
        pass
    
    elif entry == 4:
        pass

    elif entry == 5:
        pass

    elif entry == 6:
        pass

    else:
        print("Invalid Input")

    # workbook.save(filename = "sample.xlsx")
# print(spreadsheet.cell(row = 1,column = 1))
# print(spreadsheet.value)
print(spreadsheet["a:b"])
row1 = spreadsheet[1:3]
for i in row1.iterrows(min_row = 1,max_row = 10, min_col = 1,max_column = 10, values_only =True):
    print(i,end='')